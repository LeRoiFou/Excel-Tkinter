"""
Dans ce programme, on apprend à récupérer les données d'un fichier Excel comprenant 2 onglets dont chacun d'eux contient
un échantillonnage des données de la liasse fiscale

L'objectif est de récupérer dans chacun des onglets les données de la cellule A1 à la cellule D4 d'un fichier Calc de
Libreoffice converti en un fichier .xlsx d'Excel (fichier -> enregistrer sous -> Type : Excel 2007-365) et d'afficher
les données dans deux onglets distints sous la forme de champs de saisies

L'instanciation de l'objet dispose de 5 arguments :
-> L'onglet à insérer dans l'interface graphique ;
-> La plage de cellules Excel à sélectionner : cellule située en haut à gauche
-> La plage de cellules Excel à sélectionner : cellule située en bas à droite
-> Le nombre de lignes à afficher dans l'interface graphique au regard de la plage de cellules Excel sélectionnée + 1
-> Le nombre de colonnes à afficher dans l'interface graphique au regard de la plage de cellules Excel sélectionnée + 1

Éditeur : Laurent REYNAUD
Date : 06-02-2021
"""

from tkinter import *
from tkinter import ttk  # pour les onglets
from tkinter import filedialog  # pour la fenêtre d'ouverture du fichier à récupérer
from openpyxl import load_workbook  # pour récupérer les données Excel

root = Tk()
root.title('Récupération des données Excel')
root.geometry('1500x700')

"""Configuration des onglets"""
my_notebook = ttk.Notebook(root)
my_notebook.pack()

"""Fonction permettant d'ouvrir un fichier dans l'ordinateur"""
my_program = filedialog.askopenfilename()  # ouverture de la boîte de dialogue

"""Chargement de la feuille active d'Excel"""
wb = load_workbook(my_program)


class OpenProgram(Frame):
    """Classe permettant de récupérer les données du fichier Excel et de les afficher dans Python avec le module Tkinter
    """

    def __init__(self, master, number, start_cell, end_cell, row_py, column_py):
        super().__init__(master)
        self.pack()
        self.widgets(number, start_cell, end_cell, row_py, column_py)

    def widgets(self, number, start_cell, end_cell, row_py, column_py):

        """Assignation de la feuille d'Excel (feuille = worksheet = ws)"""
        ws = wb[wb.sheetnames[number]]  # number = numéro de l'onglet du Fichier Excel

        """Saisie de la plage de cellules des données à récupérer du fichier Excel"""
        range_excel = ws[start_cell:end_cell]

        """Assignation d'une liste"""
        my_list = []

        """Insertion des données Excel dans une liste"""
        for cell in range_excel:
            for x in cell:
                my_list.append(x.value)  # ajout des données de la feuille Excel dans une liste

        """Création d'un cadre principal qui s'étend sur toute la fenêtre"""
        main_frame = Frame(self)
        main_frame.pack()

        """Création d'un canvas situé à gauche du cadre principal"""
        my_canvas = Canvas(main_frame, width=1460, height=660)
        my_canvas.pack(side=LEFT)

        """Ajout d'une barre de défilement en bas du canvas sur toute sa longueur"""
        my_scrollbar_bottom = ttk.Scrollbar(main_frame, orient=VERTICAL, command=my_canvas.yview)
        my_scrollbar_bottom.pack(side=RIGHT, fill=Y)

        """Configuration du canvas"""
        my_canvas.config(yscrollcommand=my_scrollbar_bottom.set)
        my_canvas.bind('<Configure>', lambda e: my_canvas.config(scrollregion=my_canvas.bbox('all')))

        """Création d'un autre cadre à l'intérieur du canvas"""
        second_frame = Frame(my_canvas)

        """Ajout de ce nouveau cadre à la fenêtre du canvas"""
        my_canvas.create_window((0, 0), window=second_frame, anchor=NW)

        """Assignation de variables"""
        row_excel = row_py  # nombre de lignes +1
        column_excel = column_py  # nombre de colonnes +1
        nb = 0

        """Ajout des champs de saisis"""
        for i in range(1, row_excel):  # rows
            for j in range(1, column_excel):  # columns
                cell = Entry(second_frame, justify='center', font='arial 8', width=20)
                cell.grid(row=i, column=j, ipady=5)
                if my_list[nb] is None:
                    cell.insert(0, '')
                else:
                    cell.insert(0,
                                my_list[nb])  # insertion des données de la liste 'my_list' dans les champs de saisies
                nb += 1  # incrémentation +1 des données de la liste 'my_list'


"""Configuration des cadres"""
frame0 = Frame(my_notebook)
frame1 = Frame(my_notebook)

"""Ajout des onglets"""
my_notebook.add(frame0, text='Actif du bilan')
my_notebook.add(frame1, text='Passif du bilan')

open0 = OpenProgram(frame0, 7, 'A1', 'C25', 26, 4)
open1 = OpenProgram(frame1, 8, 'A1', 'D20', 21, 5)
root.mainloop()
